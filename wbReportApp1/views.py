from django.shortcuts import render
from openpyxl import load_workbook


def index(request):
    contextForTemplate = {}           #reportIndex should be for server
    return render(request, 'wbReportApp11/index.html', contextForTemplate)

def abs5(x):
    return (x*x)**0.5

def formInitialTable(wbSheet):

    temporarTable = []
    positionsList = []

    for r in range(2, wbSheet.max_row+1):
        #print(wbSheet['F'+str(loopIndex)].value)
        if wbSheet['F'+str(r)].value not in positionsList and wbSheet['F'+str(r)].value != None:
            positionsList.append(wbSheet['F'+str(r)].value)
            temporarTable.append([wbSheet['F'+str(r)].value,0,0,0,0])
    print(sorted(temporarTable))
    return sorted(temporarTable)

def processReportWB(newReportTableK, wbSheet, sales_returns):

    wbComission = 0
    logistics = 0
    logisticsPVZ = 0
    totalRealized = 0
    totalToSupplier = 0 #До вычета логистики, но после комиссии и поверенного
    dateOfReport =  wbSheet['AF2'].value
    
    #В этой части и происходит вывод данных из отчёта.
    #Когда отчёт меняется и сдвигает столбцы, тут нужно корректировки делать.
    for positionRow in newReportTableK:
        for r in range(2, wbSheet.max_row+1):
            if (wbSheet['F'+str(r)].value == positionRow[0]) and (wbSheet['K'+str(r)].value == sales_returns or (wbSheet['K'+str(r)].value == "Сторно продаж" or wbSheet['K'+str(r)].value == "Корректная продажа" and sales_returns != "Продажа")):
                positionRow[1] += round(float(wbSheet['N'+str(r)].value), 2) #Кол-во продаж позиции
                positionRow[2] += round(float(wbSheet['P'+str(r)].value), 2) #Сумма реализации позиции
                positionRow[3] += round(float(wbSheet['AC'+str(r)].value), 2) #Сумма поставщику позиции
                positionRow[4] =  round(float(positionRow[2])/float(positionRow[1]), 2) #средняя цена реализации позиции

                wbComission += float(wbSheet['AA'+str(r)].value) + float(wbSheet['AB'+str(r)].value) #Комиссия ВБ с НДС общая по неделе
                logisticsPVZ += float(wbSheet['Z'+str(r)].value) #Добавляем поверенного
                totalRealized += round(float(wbSheet['P'+str(r)].value), 2)
                totalToSupplier +=  round(float(wbSheet['AC'+str(r)].value), 2)
            elif wbSheet['F'+str(r)].value == positionRow[0] and wbSheet['K'+str(r)].value == "Логистика":
                logistics += float(wbSheet['AF'+str(r)].value) #Общая логистика без поверенным и ТК, без хранения
                
    for positionRow in newReportTableK:
        print("Отчёт по " + sales_returns)
        print("Артикул: " + str(positionRow[0]))
        print("Кол-во: " + str(positionRow[1]))
        print("Реализовано: " + str(positionRow[2]))
        print("Перевести: " + str(positionRow[3]))
        print("Ср. цена: " + str(positionRow[4]))
        print("/////////////////////////////////")
    '''
    wbComission = round(wbComission, 2)
    logistics = round(logistics, 2) 
    logisticsPVZ = round(logisticsPVZ, 2)
    totalToSupplier = round(totalToSupplier, 2) - logisticsPVZ - logistics - wbComission 
    '''
    print(wbComission)
    print(logistics)
    print(logisticsPVZ)
    print("--------")
    print(totalToSupplier)
    print("===============")
    print(totalToSupplier-logistics-logisticsPVZ - wbComission)
    
    return [newReportTableK, wbComission, logistics, logisticsPVZ, totalToSupplier, dateOfReport, totalRealized]


def processReturnMergeWB(newReportTable, newReportTableReturns):

    rowFound = False
    for returnRow  in newReportTableReturns:
        for positionRow in newReportTable:
            if positionRow[0] == returnRow[0]:
                positionRow[1] -= round(returnRow[1], 2) #Кол-во продаж позиции
                positionRow[2] -= round(returnRow[2], 2) #Сумма реализации позиции
                positionRow[3] -= round(returnRow[3], 2) #Сумма поставщику позиции
                print("Возврат: " + str(positionRow[0]))
                rowFound = True
        #WHAT if sales = 0, but return > 0? V
        if rowFound == False:
            newReportTable.append(returnRow)
            newReportTable[-1][1] *= -1
            newReportTable[-1][2] *= -1
            newReportTable[-1][3] *= -1
            
        rowFound == False

    #??
    for positionRow in newReportTable:
        if positionRow[1] == 0:
            newReportTable.remove(positionRow)


    return newReportTable

#Start method for WB Report
def parseWB(request):
    context = {}

    wbReportFile = request.FILES.get('uploadedFile', False)
    wbWorkbook = load_workbook(wbReportFile)
    wbSheet = wbWorkbook['Sheet1']


    newReportTable = formInitialTable(wbSheet)
    newReportTableReturns = formInitialTable(wbSheet)
    

    temporarDataArray = processReportWB(newReportTable, wbSheet, "Продажа")
    temporarDataArrayReturns = processReportWB(newReportTableReturns, wbSheet, "Возврат")


    #Table structure
    #    0          1               2                           3                            4
    # Article | Qty Sold | Realized price sum | Revenue after comission and Pover | Average sold price
    newReportTable = temporarDataArray[0]
    wbComission = temporarDataArray[1] #With NDS
    warehouseCost = round(float(request.POST['warehouseCost'].replace(',', '.').replace(' ', '')),2)
  #  logistics = temporarDataArray[2]+temporarDataArray[3] + warehouseCost #Logistics + Pover + input warehouse cost
   
    dateOfReport = temporarDataArray[5]

    newReportTableReturns = temporarDataArrayReturns[0]
    wbComissionWReturns = temporarDataArray[1] - temporarDataArrayReturns[1] #With NDS
    logisticsWReturns = temporarDataArray[2]+temporarDataArray[3] - temporarDataArrayReturns[3] #Logistics + Pover
    
    totalRealizedWReturns = temporarDataArray[6] - temporarDataArrayReturns[6]
    totalToSupplierWReturns = totalRealizedWReturns - logisticsWReturns - wbComissionWReturns - warehouseCost

    contextForTemplate = {
        'salesList': newReportTable,
        'returnsList': newReportTableReturns,
        'mergedList': processReturnMergeWB(newReportTable, newReportTableReturns),
        'wbComissionWReturns': wbComissionWReturns,
        'logisticsWReturns': temporarDataArray[2],
        'pover': round(temporarDataArray[3] - temporarDataArrayReturns[3], 2),
        'warehouseCost': warehouseCost,
        'dateOfReport': dateOfReport,
        'totalRealizedWReturns': totalRealizedWReturns,
        'totalToSupplierWReturns': totalToSupplierWReturns,
        
    }

    '''
    for positionElementRow in processReturnMergeWB(newReportTable, newReportTableReturns):
                
        dtSheet.insert_rows(maxRow, 1)
        dtSheet['A'+str(maxRow)].value = dateOfReport
        #dtSheet['A'+str(maxRow)]._style = dtSheet['A'+str(maxRow-2)]._style

        dtSheet['B'+str(maxRow)].value = "Вайлдберриз"
        dtSheet['C'+str(maxRow)].value = positionElementRow[1] #Кол-во
        dtSheet['D'+str(maxRow)].value = logisticsWReturns  if firstPositionRow else 0
        dtSheet['E'+str(maxRow)].value = positionElementRow[2] #Realized
        dtSheet['F'+str(maxRow)].value = positionElementRow[3] #- wbComission
        dtSheet['G'+str(maxRow)].value = "=F"+str(maxRow)+"-D"+str(maxRow)
        dtSheet['H'+str(maxRow)].value = "=G"+str(maxRow)+"-M"+str(maxRow)+"*C"+str(maxRow)
        dtSheet['I'+str(maxRow)].value = "=H"+str(maxRow)+"-J"+str(maxRow)
        dtSheet['J'+str(maxRow)].value = "=G"+str(maxRow)+"*0.15" if tax15DictWB.get(positionElementRow[0]) else "=H"+str(maxRow)+"*0.15"
        dtSheet['K'+str(maxRow)].value = "-"
        dtSheet['L'+str(maxRow)].value = nameDictWB.get(positionElementRow[0])
        dtSheet['M'+str(maxRow)].value = ssDictWB.get(positionElementRow[0])
        dtSheet['N'+str(maxRow)].value = "=F"+str(maxRow)+"/C"+str(maxRow)

        dtSheet.row_dimensions[maxRow].height = 41

        for i in range(0,14):
            dtSheet[maxRow][i]._style = dtSheet[maxRow-1][i]._style

        firstPositionRow = False
        maxRow += 1
    '''
    #reportResult must be template
    return render(request, 'wbReportApp11/results.html', contextForTemplate)